from __future__ import annotations

import base64
import json
import mimetypes
from pathlib import Path
from typing import Any

import structlog

from domain.interfaces import OcrPort

logger = structlog.get_logger(__name__)

IMAGE_EXTS = (".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp")
PDF_EXTS = (".pdf",)
EXCEL_EXTS = (".xlsx", ".xls")
TEXT_EXTS = (".txt", ".log", ".csv", ".tsv", ".json", ".xml", ".yaml", ".yml", ".conf", ".ini", ".cfg")
DOC_EXTS = (".doc", ".docx", ".rtf")

ALL_SUPPORTED = IMAGE_EXTS + PDF_EXTS + EXCEL_EXTS + TEXT_EXTS + DOC_EXTS


class MockOcrProcessor(OcrPort):
    async def extract_text(self, attachment_url: str) -> str:
        lower = attachment_url.lower()
        if "bad" in lower or "fail" in lower:
            raise RuntimeError("OCR failed: unreadable image")
        if lower.endswith(EXCEL_EXTS):
            return "excel:sheet1:username=APP_X"
        if any(lower.endswith(ext) for ext in IMAGE_EXTS):
            return "ocr: Create user from screenshot"
        if lower.endswith(PDF_EXTS):
            return "pdf: Create user APP_PDF in DEVDB, grant Read Only role."
        if lower.endswith(TEXT_EXTS):
            return "log: 2026-08-04 ERROR ORA-28000 account APP_LOCKED is locked on PRODDB"
        if lower.endswith(DOC_EXTS):
            return "doc: Request to create user SVC_REPORT in DEVDB with DBA role."
        logger.warning("ocr_skip_unsupported", url_suffix=lower[-12:])
        return ""


class OcrProcessor(OcrPort):
    """Extracts text from any attachment type using Claude via AWS Bedrock.

    Supports: images (vision), PDFs (document), logs/text (direct read),
    Excel (parsed to text), and docs (sent to Claude for summarization).
    """

    def __init__(self, model_id: str, region: str) -> None:
        import boto3

        self._model_id = model_id
        self._client = boto3.client("bedrock-runtime", region_name=region)

    async def extract_text(self, attachment_url: str) -> str:
        lower = attachment_url.lower()

        if any(lower.endswith(ext) for ext in IMAGE_EXTS):
            return await self._extract_from_image(attachment_url)

        if any(lower.endswith(ext) for ext in PDF_EXTS):
            return await self._extract_from_pdf(attachment_url)

        if any(lower.endswith(ext) for ext in TEXT_EXTS):
            return await self._extract_from_text_file(attachment_url)

        if any(lower.endswith(ext) for ext in EXCEL_EXTS):
            return await self._extract_from_excel(attachment_url)

        if any(lower.endswith(ext) for ext in DOC_EXTS):
            return await self._extract_from_doc(attachment_url)

        logger.warning("ocr_skip_unsupported", url=lower[-30:])
        return ""

    # --- Image: send to Claude Vision ---

    async def _extract_from_image(self, url: str) -> str:
        image_bytes = await self._fetch_file(url)
        media_type = self._guess_image_media_type(url)

        content: list[dict[str, Any]] = [
            {
                "type": "image",
                "source": {
                    "type": "base64",
                    "media_type": media_type,
                    "data": base64.b64encode(image_bytes).decode("utf-8"),
                },
            },
            {
                "type": "text",
                "text": (
                    "Extract ALL text from this image. Include every piece of information: "
                    "usernames, database names, roles, hostnames, environments, form fields, "
                    "table data, error messages, logs — everything. "
                    "Return the extracted content as plain text, preserving structure."
                ),
            },
        ]

        logger.info("ocr_image_invoke", model=self._model_id, bytes=len(image_bytes))
        return await self._invoke_claude(content)

    # --- PDF: extract text, send to Claude for structured extraction ---

    async def _extract_from_pdf(self, url: str) -> str:
        pdf_bytes = await self._fetch_file(url)
        text = self._pdf_to_text(pdf_bytes)

        if not text.strip():
            # Scanned PDF with no selectable text — send pages as images to Claude
            return await self._ocr_scanned_pdf(pdf_bytes)

        if len(text) <= 8000:
            return text

        # Large PDF — ask Claude to extract the relevant request details
        return await self._summarize_with_claude(
            text,
            "Extract all IT service request details from this PDF: "
            "usernames, databases, roles, hostnames, environments, actions requested. "
            "Return structured plain text with all relevant fields.",
        )

    def _pdf_to_text(self, pdf_bytes: bytes) -> str:
        try:
            import PyPDF2
            import io

            reader = PyPDF2.PdfReader(io.BytesIO(pdf_bytes))
            pages = []
            for page in reader.pages:
                page_text = page.extract_text() or ""
                if page_text.strip():
                    pages.append(page_text.strip())
            return "\n\n".join(pages)
        except Exception as e:
            logger.warning("pdf_text_extraction_failed", error=str(e))
            return ""

    async def _ocr_scanned_pdf(self, pdf_bytes: bytes) -> str:
        """For scanned PDFs without selectable text, convert pages to images and OCR via Claude."""
        try:
            import PyPDF2
            import io

            reader = PyPDF2.PdfReader(io.BytesIO(pdf_bytes))
            # Send the raw PDF bytes as a document to Claude (Bedrock supports PDF)
            content: list[dict[str, Any]] = [
                {
                    "type": "document",
                    "source": {
                        "type": "base64",
                        "media_type": "application/pdf",
                        "data": base64.b64encode(pdf_bytes).decode("utf-8"),
                    },
                },
                {
                    "type": "text",
                    "text": (
                        "Extract ALL text from this PDF document. Include every piece of "
                        "information: usernames, database names, roles, hostnames, environments, "
                        "form fields, table data, error messages — everything visible. "
                        "Return as plain text preserving structure."
                    ),
                },
            ]
            return await self._invoke_claude(content)
        except Exception as e:
            logger.warning("scanned_pdf_ocr_failed", error=str(e))
            return ""

    # --- Text/Log/CSV files: read directly ---

    async def _extract_from_text_file(self, url: str) -> str:
        file_bytes = await self._fetch_file(url)

        for encoding in ("utf-8", "latin-1", "cp1252"):
            try:
                text = file_bytes.decode(encoding)
                break
            except (UnicodeDecodeError, ValueError):
                continue
        else:
            text = file_bytes.decode("utf-8", errors="replace")

        # Truncate very large files but keep enough context
        max_chars = 15000
        if len(text) > max_chars:
            logger.info("text_file_truncated", original=len(text), kept=max_chars)
            text = text[:max_chars] + "\n\n[... truncated ...]"

        logger.info("text_file_read", chars=len(text), url=url[-30:])
        return text

    # --- Excel: parse sheets to text ---

    async def _extract_from_excel(self, url: str) -> str:
        file_bytes = await self._fetch_file(url)

        try:
            import openpyxl
            import io

            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
            parts = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = []
                for row in ws.iter_rows(values_only=True):
                    cells = [str(c) if c is not None else "" for c in row]
                    if any(cells):
                        rows.append("\t".join(cells))
                if rows:
                    parts.append(f"[Sheet: {sheet_name}]\n" + "\n".join(rows))
            wb.close()
            text = "\n\n".join(parts)
            logger.info("excel_parsed", sheets=len(parts), chars=len(text))
            return text
        except Exception as e:
            logger.warning("excel_parse_failed", error=str(e))
            # Fallback: send to Claude as document
            return await self._summarize_with_claude(
                f"[Excel file could not be parsed locally: {e}]",
                "The attached file is an Excel spreadsheet. Extract any relevant data.",
            )

    # --- Word docs: extract text ---

    async def _extract_from_doc(self, url: str) -> str:
        file_bytes = await self._fetch_file(url)

        if url.lower().endswith(".docx"):
            return self._docx_to_text(file_bytes)

        # .doc / .rtf — send to Claude for extraction
        content: list[dict[str, Any]] = [
            {
                "type": "document",
                "source": {
                    "type": "base64",
                    "media_type": "application/msword",
                    "data": base64.b64encode(file_bytes).decode("utf-8"),
                },
            },
            {
                "type": "text",
                "text": "Extract all text from this document. Return as plain text.",
            },
        ]
        return await self._invoke_claude(content)

    def _docx_to_text(self, docx_bytes: bytes) -> str:
        try:
            import zipfile
            import io
            import xml.etree.ElementTree as ET

            with zipfile.ZipFile(io.BytesIO(docx_bytes)) as z:
                with z.open("word/document.xml") as f:
                    tree = ET.parse(f)
            ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
            paragraphs = []
            for p in tree.iter(f"{{{ns['w']}}}p"):
                texts = [t.text for t in p.iter(f"{{{ns['w']}}}t") if t.text]
                if texts:
                    paragraphs.append("".join(texts))
            return "\n".join(paragraphs)
        except Exception as e:
            logger.warning("docx_parse_failed", error=str(e))
            return ""

    # --- Shared helpers ---

    async def _invoke_claude(self, content: list[dict[str, Any]]) -> str:
        body = {
            "anthropic_version": "bedrock-2023-05-31",
            "max_tokens": 4096,
            "messages": [{"role": "user", "content": content}],
        }

        response = self._client.invoke_model(
            modelId=self._model_id,
            contentType="application/json",
            accept="application/json",
            body=json.dumps(body),
        )

        payload = json.loads(response["body"].read())
        result_content = payload.get("content") or []
        if not result_content or not isinstance(result_content[0], dict) or "text" not in result_content[0]:
            raise ValueError("Empty response from Claude")

        extracted = result_content[0]["text"]
        logger.info("claude_extraction_complete", chars=len(extracted))
        return extracted

    async def _summarize_with_claude(self, text: str, instruction: str) -> str:
        content: list[dict[str, Any]] = [
            {"type": "text", "text": f"{instruction}\n\n---\n\n{text}"},
        ]
        return await self._invoke_claude(content)

    async def _fetch_file(self, attachment_url: str) -> bytes:
        path = Path(attachment_url)
        if path.is_file():
            return path.read_bytes()

        import httpx

        async with httpx.AsyncClient(timeout=60.0) as client:
            resp = await client.get(attachment_url)
            resp.raise_for_status()
            return resp.content

    def _guess_image_media_type(self, url: str) -> str:
        mime, _ = mimetypes.guess_type(url)
        if mime and mime.startswith("image/"):
            return mime
        lower = url.lower()
        if lower.endswith(".png"):
            return "image/png"
        if lower.endswith((".jpg", ".jpeg")):
            return "image/jpeg"
        if lower.endswith(".gif"):
            return "image/gif"
        if lower.endswith(".webp"):
            return "image/webp"
        return "image/png"
